msg = "Hello World"
print(msg)
import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from contextlib import suppress
import time
import os
from webdriver_manager.firefox import GeckoDriverManager
from openpyxl import load_workbook

driver = webdriver.Firefox(executable_path=GeckoDriverManager().install())

#find the row in the worksheet
wb=load_workbook(filename='D:\Linux Ubuntu 2019\Documents\Marktplaats verkopen.xlsx')
ws = wb.active
r=''
for x in range(len(ws['B'])):
    if ws['B'][x].value=='Super Nintendo cartridge opbergkoffer Donkey Kong':
        #r=x
        break

# x is de regel van de fotos die je wil uploaden
x = 1

file = open('D:\Linux Ubuntu 2019\Documents\marktplaats.txt', 'r')

#inlogpagina
driver.get('https://www.marktplaats.nl/account/login.html')
driver.find_element_by_name("j_username").send_keys(file.readline())
driver.find_element_by_name("j_password").send_keys(file.readline())

driver.find_element_by_id("account-login-button").click()
driver.implicitly_wait(3)
driver.find_element_by_id("account-login-button").click()

driver.get('https://www.marktplaats.nl/plaats')

#strings uit CSV
category_str=ws['A'][x].value
titel_str=ws['B'][x].value
beschrijving_str=ws['C'][x].value
conditie_str=ws['D'][x].value
prijstype_str=ws['E'][x].value
vraagprijs_str=str(ws['F'][x].value)
biedenvanaf_str=str(ws['G'][x].value)
ophalenofverzenden_str=ws['H'][x].value
fotolocatie_str=ws['I'][x].value
gewicht_str=ws['J'][x].value
type_str=ws['R'][x].value
brievenbus_str=ws['S'][x].value
foto_list = os.listdir(fotolocatie_str)
foto_list.sort()

def click(xpath):
    try:
        element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, xpath)))
        return driver.find_element_by_xpath(xpath).click()
    except Exception as e:
        print(f'Got exception of type {type(e)}: {e}')

def sendkeys(xpath,inputstring):
    try:
        element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, xpath)))
        return driver.find_element_by_xpath(xpath).send_keys(inputstring)
    except Exception as e:
        print(f'Got exception of type {type(e)}: {e}')

wait = WebDriverWait(driver, 5)
element = wait.until(EC.element_to_be_clickable((By.ID, 'category-keywords')))
driver.find_element_by_id('category-keywords').send_keys(category_str)

element = wait.until(EC.element_to_be_clickable((By.ID, 'find-category')))
driver.find_element_by_id('find-category').click()

element = wait.until(EC.element_to_be_clickable((By.ID, 'category-selection-submit')))
driver.find_element_by_id('category-selection-submit').click()

#foto's uploaden, bij foto 1 klikken dat je geen bonus wil
driver.find_element_by_xpath('//*[@id="uploader-container-0"]/div[2]/input[1]').send_keys(fotolocatie_str+'\\'+foto_list[0])

for i in range(1,len(foto_list)):
    element = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '//*[@id="uploader-container-'+str(i)+'"]/div[2]/input[1]')))
    driver.find_element_by_xpath('//*[@id="uploader-container-'+str(i)+'"]/div[2]/input[1]').send_keys(fotolocatie_str+'\\'+foto_list[i])

element = wait.until(EC.element_to_be_clickable((By.ID, 'title')))
driver.find_element_by_id('title').clear()
driver.find_element_by_id('title').send_keys(titel_str)

element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, """//div[@id='mceu_11']/iframe[1]""")))
driver.find_element_by_xpath("""//div[@id='mceu_11']/iframe[1]""").click()
driver.find_element_by_xpath("""//div[@id='mceu_11']/iframe[1]""").send_keys(beschrijving_str)

# pas de Conditie aan
sendkeys("//select[contains(@name,'singleSelectAttribute[condition]')]",conditie_str)


# click Type Bouwen
if type_str == 'Bouwen':
    click("//input[contains(@value,'Bouwen')]")
elif type_str == 'Complete set':
    click("//select[@name='singleSelectAttribute[type]']")
    sendkeys("//select[@name='singleSelectAttribute[type]']",type_str)
elif type_str == '(Natuur)geluiden':
    sendkeys("//select[@name='singleSelectAttribute[type]']",type_str)
    

element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, """//select[@name='price.typeValue']""")))
driver.find_element_by_xpath("""//select[@name='price.typeValue']""").send_keys(prijstype_str)

element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, """//div[@id='syi-bidding-price']//input[@placeholder='0,00']""")))
driver.find_element_by_xpath("""//div[@id='syi-bidding-price']//input[@placeholder='0,00']""").send_keys(vraagprijs_str)

element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, """//div[@id='syi-bidding-minimumprice']//input[@placeholder='0,00']""")))
driver.find_element_by_xpath("""//div[@id='syi-bidding-minimumprice']//input[@placeholder='0,00']""").send_keys(biedenvanaf_str)

element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, """//select[@name='attribute[delivery]']""")))
driver.find_element_by_xpath("""//select[@name='attribute[delivery]']""").send_keys(ophalenofverzenden_str)

element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, """//div[@id='feature-FREE']""")))
driver.find_element_by_xpath("""//div[@id='feature-FREE']""").click()

element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, """//*[@id="shippingMethod0"]""")))
driver.find_element_by_xpath("""//*[@id="shippingMethod0"]""").click()

if brievenbus_str.lower() == 'ja':
    click("(//input[@name='fitsInMailBox'])[1]")
    if gewicht_str <= 20:
        click("//input[contains(@value,'1000_letters_10')]")
    elif gewicht_str <= 50:
        click("(//input[contains(@name,'shippingDimension')])[4]")
    elif gewicht_str <= 100:
        click("//input[contains(@value,'1000_letters_75')]")
    elif gewicht_str <= 350:
        click("//input[contains(@value,'1000_letters_175')]")
    else:
        click("//input[contains(@value,'1018_parcels_1000')]")

    # click op de Opslaan knop om het gewicht op te slaan
    click("//button[@class='mp-Button mp-Button--postnl'][contains(.,'Opslaan')]")
    
elif brievenbus_str.lower() == 'nee': 
    click("(//input[@name='fitsInMailBox'])[2]")
    if gewicht_str <= 10000:
        click("""//*[@id="3000_parcels_5000"]""")
    else:
        click("""//*[@id="3001_parcels_20000"]""")

    # click op de Opslaan knop om het gewicht op te slaan
    click("//button[@class='mp-Button mp-Button--postnl'][contains(.,'Opslaan')]")
else:
    print(f'de waarde [{brievenbus_str}] wordt niet herkend')

#plaatst de daadwerkelijke advertentie
element = wait.until(EC.element_to_be_clickable((By.ID, 'syi-place-ad-button')))
driver.find_element_by_id('syi-place-ad-button').click()
    
